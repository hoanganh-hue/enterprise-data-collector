"""
Basic Excel Exporter for Enterprise Data Collector
Basic Excel export functionality
Author: MiniMax Agent
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """
    Basic Excel exporter cho company data
    """
    
    def __init__(self, output_dir: str = "Outputs"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger(__name__)
    
    def export_companies(
        self,
        companies: List[Dict[str, Any]],
        filename: Optional[str] = None,
        sheet_name: str = "Companies"
    ) -> str:
        """
        Export danh sách công ty ra Excel
        
        Args:
            companies: Danh sách công ty
            filename: Tên file (auto-generate nếu None)
            sheet_name: Tên sheet
            
        Returns:
            Đường dẫn file đã tạo
        """
        
        if not companies:
            raise ValueError("No companies to export")
        
        # Auto-generate filename
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"companies_export_{timestamp}.xlsx"
        
        output_path = self.output_dir / filename
        
        try:
            # Tạo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Lấy headers từ company đầu tiên
            headers = list(companies[0].keys())
            
            # Viết headers
            self._write_headers(ws, headers)
            
            # Viết dữ liệu
            for row_idx, company in enumerate(companies, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    value = company.get(header, '')
                    
                    # Xử lý giá trị đặc biệt
                    if value is None:
                        value = ''
                    elif isinstance(value, (list, dict)):
                        value = str(value)
                    
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-resize columns
            self._auto_resize_columns(ws)
            
            # Apply basic formatting
            self._apply_basic_formatting(ws, len(companies) + 1)
            
            # Lưu file
            wb.save(output_path)
            
            self.logger.info(f"Exported {len(companies)} companies to {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Failed to export Excel: {e}")
            raise
    
    def _write_headers(self, ws, headers: List[str]):
        """
        Viết headers với formatting
        """
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _auto_resize_columns(self, ws):
        """
        Tự động điều chỉnh độ rộng cột
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Đặt độ rộng (giới hạn 50 ký tự)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_basic_formatting(self, ws, total_rows: int):
        """
        Áp dụng formatting cơ bản
        """
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Áp dụng border cho tất cả các cell có dữ liệu
        for row in ws.iter_rows(min_row=1, max_row=total_rows):
            for cell in row:
                cell.border = thin_border
        
        # Freeze first row
        ws.freeze_panes = "A2"
    
    def export_summary_report(
        self,
        companies: List[Dict[str, Any]],
        stats: Dict[str, Any],
        filename: Optional[str] = None
    ) -> str:
        """
        Export báo cáo tổng hợp
        
        Args:
            companies: Danh sách công ty
            stats: Thống kê
            filename: Tên file
            
        Returns:
            Đường dẫn file đã tạo
        """
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"summary_report_{timestamp}.xlsx"
        
        output_path = self.output_dir / filename
        
        try:
            wb = openpyxl.Workbook()
            
            # Sheet 1: Summary Statistics
            ws_summary = wb.active
            ws_summary.title = "Summary"
            self._create_summary_sheet(ws_summary, stats)
            
            # Sheet 2: Company Data
            ws_companies = wb.create_sheet("Companies")
            if companies:
                headers = list(companies[0].keys())
                self._write_headers(ws_companies, headers)
                
                for row_idx, company in enumerate(companies, start=2):
                    for col_idx, header in enumerate(headers, start=1):
                        value = company.get(header, '')
                        if value is None:
                            value = ''
                        elif isinstance(value, (list, dict)):
                            value = str(value)
                        ws_companies.cell(row=row_idx, column=col_idx, value=value)
                
                self._auto_resize_columns(ws_companies)
                self._apply_basic_formatting(ws_companies, len(companies) + 1)
            
            wb.save(output_path)
            
            self.logger.info(f"Exported summary report to {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Failed to export summary report: {e}")
            raise
    
    def _create_summary_sheet(self, ws, stats: Dict[str, Any]):
        """
        Tạo sheet thống kê tổng hợp
        """
        # Title
        ws['A1'] = "Báo cáo Tổng hợp Thu thập Dữ liệu Doanh nghiệp"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Timestamp
        ws['A2'] = f"Tạo lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=12)
        
        # Basic stats
        row = 4
        ws[f'A{row}'] = "Thống kê tổng quát:"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        for key, value in stats.items():
            if isinstance(value, dict):
                ws[f'A{row}'] = f"{key}:"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for sub_key, sub_value in value.items():
                    ws[f'B{row}'] = f"  {sub_key}: {sub_value}"
                    row += 1
            else:
                ws[f'A{row}'] = f"{key}: {value}"
                row += 1
        
        # Auto-resize
        self._auto_resize_columns(ws)