"""
Enhanced Excel Exporter for Enterprise Data Collector
31-column enhanced Excel export with dual data source integration
Author: MiniMax Agent
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

from ..models.enhanced_company import EnhancedCompany


class EnhancedExcelExporter:
    """
    Enhanced Excel exporter với 31 cột và formatting chuyên nghiệp
    """
    
    def __init__(self, output_dir: str = "Outputs"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger(__name__)
    
    def export_enhanced_companies(
        self,
        companies: List[EnhancedCompany],
        filename: Optional[str] = None,
        include_charts: bool = True
    ) -> str:
        """
        Export enhanced companies với 31 cột và formatting chuyên nghiệp
        
        Args:
            companies: Danh sách EnhancedCompany
            filename: Tên file (auto-generate nếu None)
            include_charts: Có bao gồm biểu đồ không
            
        Returns:
            Đường dẫn file đã tạo
        """
        
        if not companies:
            raise ValueError("No companies to export")
        
        # Auto-generate filename
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"enhanced_companies_{timestamp}.xlsx"
        
        output_path = self.output_dir / filename
        
        try:
            # Tạo workbook với styles
            wb = openpyxl.Workbook()
            self._setup_styles(wb)
            
            # Sheet 1: Main Data (31 columns)
            ws_main = wb.active
            ws_main.title = "Company Data"
            self._create_main_data_sheet(ws_main, companies)
            
            # Sheet 2: Summary & Statistics
            ws_summary = wb.create_sheet("Summary")
            self._create_enhanced_summary_sheet(ws_summary, companies)
            
            # Sheet 3: Data Source Analysis
            ws_sources = wb.create_sheet("Data Sources")
            self._create_data_source_analysis(ws_sources, companies)
            
            # Thêm charts nếu yêu cầu
            if include_charts:
                self._add_charts(ws_summary, companies)
            
            # Lưu file
            wb.save(output_path)
            
            self.logger.info(f"Exported {len(companies)} enhanced companies to {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Failed to export enhanced Excel: {e}")
            raise
    
    def _setup_styles(self, wb):
        """
        Thiết lập các styles cho workbook
        """
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, color="FFFFFF", size=12)
        header_style.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_style.border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        # Data style
        data_style = NamedStyle(name="data")
        data_style.alignment = Alignment(vertical="center", wrap_text=True)
        data_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Important data style (for key fields)
        important_style = NamedStyle(name="important")
        important_style.font = Font(bold=True)
        important_style.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        important_style.alignment = Alignment(vertical="center", wrap_text=True)
        important_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Thêm styles vào workbook
        wb.add_named_style(header_style)
        wb.add_named_style(data_style)
        wb.add_named_style(important_style)
    
    def _create_main_data_sheet(self, ws, companies: List[EnhancedCompany]):
        """
        Tạo sheet dữ liệu chính với 31 cột
        """
        # Lấy headers (31 cột)
        headers = EnhancedCompany.get_excel_headers()
        
        # Viết headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.style = "header"
        
        # Viết dữ liệu
        for row_idx, company in enumerate(companies, start=2):
            row_data = company.to_excel_row()
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Áp dụng style dựa trên cột
                if col_idx in [1, 2, 4, 5]:  # Các cột quan trọng: MST, Tên, Đại diện, ĐT
                    cell.style = "important"
                else:
                    cell.style = "data"
        
        # Đặt độ cao hàng header
        ws.row_dimensions[1].height = 40
        
        # Auto-resize columns với giới hạn thông minh
        self._smart_resize_columns(ws)
        
        # Freeze panes
        ws.freeze_panes = "D2"  # Freeze 3 cột đầu và header
        
        # Thêm filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(companies) + 1}"
        
        # Data validation cho một số cột
        self._add_data_validation(ws, len(companies) + 1)
    
    def _smart_resize_columns(self, ws):
        """
        Tự động điều chỉnh độ rộng cột thông minh
        """
        # Độ rộng đặc biệt cho từng loại cột
        special_widths = {
            1: 15,   # Mã số thuế
            2: 30,   # Tên công ty
            3: 40,   # Địa chỉ
            4: 20,   # Người đại diện
            5: 15,   # Điện thoại
            8: 15,   # Tỉnh thành phố
            17: 35,  # Ngành nghề chính
        }
        
        for column in ws.columns:
            column_letter = get_column_letter(column[0].column)
            col_idx = column[0].column
            
            if col_idx in special_widths:
                ws.column_dimensions[column_letter].width = special_widths[col_idx]
            else:
                # Tính toán độ rộng dựa trên nội dung
                max_length = 0
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                # Đặt độ rộng (giới hạn 12-40)
                adjusted_width = max(12, min(max_length + 2, 40))
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_data_validation(self, ws, total_rows: int):
        """
        Thêm data validation cho một số cột
        """
        # Validation cho cột tình trạng hoạt động
        status_validation = DataValidation(
            type="list",
            formula1='"Hoạt động,Ngừng hoạt động,Chưa hoạt động"',
            allow_blank=True
        )
        
        # Áp dụng cho cột tình trạng (cột 20)
        status_validation.add(f"T2:T{total_rows}")
        ws.add_data_validation(status_validation)
    
    def _create_enhanced_summary_sheet(self, ws, companies: List[EnhancedCompany]):
        """
        Tạo sheet tổng hợp nâng cao
        """
        # Title
        ws['A1'] = "Báo cáo Tổng hợp Thu thập Dữ liệu Doanh nghiệp (Enhanced)"
        ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        ws.merge_cells('A1:E1')
        
        # Timestamp và thông tin cơ bản
        ws['A3'] = f"Tạo lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Tổng số công ty: {len(companies)}"
        
        # Thống kê theo nguồn dữ liệu
        row = 6
        ws[f'A{row}'] = "Thống kê theo nguồn dữ liệu:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        data_source_stats = {}
        for company in companies:
            source = company.data_source
            data_source_stats[source] = data_source_stats.get(source, 0) + 1
        
        row += 1
        for source, count in data_source_stats.items():
            ws[f'B{row}'] = f"{source}: {count} công ty ({count/len(companies)*100:.1f}%)"
            row += 1
        
        # Thống kê theo tình trạng
        row += 1
        ws[f'A{row}'] = "Thống kê theo tình trạng hoạt động:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        status_stats = {}
        for company in companies:
            status = company.tinh_trang_hoat_dong or "Không xác định"
            status_stats[status] = status_stats.get(status, 0) + 1
        
        row += 1
        for status, count in status_stats.items():
            ws[f'B{row}'] = f"{status}: {count} công ty"
            row += 1
        
        # Thống kê theo tỉnh thành
        row += 1
        ws[f'A{row}'] = "Top 10 tỉnh thành có nhiều công ty nhất:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        province_stats = {}
        for company in companies:
            province = company.tinh_thanh_pho or "Không xác định"
            province_stats[province] = province_stats.get(province, 0) + 1
        
        # Sắp xếp và lấy top 10
        top_provinces = sorted(province_stats.items(), key=lambda x: x[1], reverse=True)[:10]
        
        row += 1
        for i, (province, count) in enumerate(top_provinces, 1):
            ws[f'B{row}'] = f"{i}. {province}: {count} công ty"
            row += 1
        
        # Chất lượng dữ liệu
        row += 1
        ws[f'A{row}'] = "Chất lượng dữ liệu:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        # Đếm các trường có dữ liệu
        field_completeness = self._calculate_field_completeness(companies)
        
        row += 1
        for field, percentage in field_completeness.items():
            ws[f'B{row}'] = f"{field}: {percentage:.1f}% hoàn chỉnh"
            row += 1
    
    def _calculate_field_completeness(self, companies: List[EnhancedCompany]) -> Dict[str, float]:
        """
        Tính toán tỷ lệ hoàn chỉnh cho các trường quan trọng
        """
        if not companies:
            return {}
        
        important_fields = {
            'Tên công ty': 'ten_cong_ty',
            'Địa chỉ': 'dia_chi_dang_ky',
            'Người đại diện': 'nguoi_dai_dien',
            'Điện thoại': 'dien_thoai',
            'Ngành nghề': 'nganh_nghe_kinh_doanh_chinh',
            'Tỉnh thành': 'tinh_thanh_pho'
        }
        
        completeness = {}
        total_companies = len(companies)
        
        for display_name, field_name in important_fields.items():
            complete_count = 0
            for company in companies:
                value = getattr(company, field_name, None)
                if value and str(value).strip():
                    complete_count += 1
            
            completeness[display_name] = (complete_count / total_companies) * 100
        
        return completeness
    
    def _create_data_source_analysis(self, ws, companies: List[EnhancedCompany]):
        """
        Tạo sheet phân tích nguồn dữ liệu
        """
        ws['A1'] = "Phân tích Nguồn Dữ liệu"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Phân tích chi tiết theo nguồn
        api_only = [c for c in companies if c.data_source == "api"]
        hsctvn_only = [c for c in companies if c.data_source == "hsctvn"]
        dual_source = [c for c in companies if c.data_source == "dual"]
        
        row = 3
        ws[f'A{row}'] = "Phân phối nguồn dữ liệu:"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'B{row}'] = f"Chỉ từ API chính: {len(api_only)} công ty"
        row += 1
        ws[f'B{row}'] = f"Chỉ từ HSCTVN: {len(hsctvn_only)} công ty"
        row += 1
        ws[f'B{row}'] = f"Kết hợp 2 nguồn: {len(dual_source)} công ty"
        
        # Lợi ích của việc kết hợp
        row += 2
        ws[f'A{row}'] = "Lợi ích của việc tích hợp HSCTVN:"
        ws[f'A{row}'].font = Font(bold=True)
        
        # Đếm số công ty có thêm thông tin từ HSCTVN
        enhanced_phone = len([c for c in dual_source if c.dien_thoai_dai_dien])
        enhanced_legal_rep = len([c for c in dual_source if c.dai_dien_phap_luat])
        enhanced_address = len([c for c in dual_source if c.dia_chi_thue])
        
        row += 1
        ws[f'B{row}'] = f"Có thêm số điện thoại: {enhanced_phone} công ty"
        row += 1
        ws[f'B{row}'] = f"Có thêm đại diện pháp luật: {enhanced_legal_rep} công ty"
        row += 1
        ws[f'B{row}'] = f"Có thêm địa chỉ thuế: {enhanced_address} công ty"
    
    def _add_charts(self, ws, companies: List[EnhancedCompany]):
        """
        Thêm biểu đồ vào sheet
        """
        try:
            # Biểu đồ phân phối theo tình trạng
            status_stats = {}
            for company in companies:
                status = company.tinh_trang_hoat_dong or "Không xác định"
                status_stats[status] = status_stats.get(status, 0) + 1
            
            # Tạo dữ liệu cho chart
            start_row = 25
            ws[f'G{start_row}'] = "Tình trạng"
            ws[f'H{start_row}'] = "Số lượng"
            
            for i, (status, count) in enumerate(status_stats.items(), 1):
                ws[f'G{start_row + i}'] = status
                ws[f'H{start_row + i}'] = count
            
            # Tạo chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Phân phối theo Tình trạng Hoạt động"
            chart.y_axis.title = 'Số lượng công ty'
            chart.x_axis.title = 'Tình trạng'
            
            data = Reference(ws, min_col=8, min_row=start_row, max_row=start_row + len(status_stats))
            cats = Reference(ws, min_col=7, min_row=start_row + 1, max_row=start_row + len(status_stats))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "J25")
            
        except Exception as e:
            self.logger.warning(f"Failed to add charts: {e}")